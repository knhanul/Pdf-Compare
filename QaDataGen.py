import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from datetime import datetime
import os

def process_excel_v4():
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title="엑셀 파일을 선택하세요",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    
    if not file_path: return

    try:
        # data_only=True를 사용해 수식이 아닌 계산된 값을 가져옴
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        new_ws = wb.create_sheet("검증결과_최종정리_완료")
        headers = ["제목", "검증내용", "테스트케이스", "기간", "시작일", "종료일", "담당"]
        new_ws.append(headers)

        # 2026년 1월 기준 날짜 설정
        today = datetime.now()
        current_yyyy = today.year
        current_mm = int(today.strftime('%m'))
        today_ym = current_yyyy * 100 + current_mm # 202601
        
        # 1월이면 기준 연도를 전년도(2025)로 설정
        base_year = current_yyyy - 1 if current_mm == 1 else current_yyyy

        last_assignee = ""
        row_groups = []
        current_group = []

        # 1. 오직 '테두리'만을 기준으로 그룹화 (텍스트 존재 여부 무시)
        for row in ws.iter_rows(min_row=2):
            # 행에 실질적인 데이터가 아예 없는 경우는 건너뜀
            if not any(cell.value for cell in row): continue
            
            # A, B, C, D열 중 하나라도 상단 테두리가 있으면 새로운 박스의 시작
            # (담당자 E열은 테두리가 없으므로 제외하고 체크)
            has_top_border = any(row[i].border.top.style is not None for i in range(4))

            if has_top_border and current_group:
                row_groups.append(current_group)
                current_group = []
            
            current_group.append(row)
        
        if current_group:
            row_groups.append(current_group)

        # 2. 그룹별 데이터 병합 및 상태 판별
        for group in row_groups:
            title_parts = []
            verification_parts = []
            tc_val = ""
            raw_period = ""
            
            for row in group:
                # 제목: 행이 나뉘어 있으면 '공백 하나'를 추가해서 머지
                if row[0].value: 
                    title_parts.append(str(row[0].value).strip())
                # 검증내용: 행별로 엔터(\n)를 넣어 줄바꿈 머지
                if row[1].value: 
                    verification_parts.append(str(row[1].value).strip())
                # 테스트케이스 (그룹 내 첫 번째 값 유지)
                if not tc_val and row[2].value: 
                    tc_val = str(row[2].value)
                # 기간 (흩어진 텍스트 결합)
                if row[3].value: 
                    raw_period += str(row[3].value).replace(" ", "")
                # 담당자 (값이 나올 때까지 이전 담당자 유지)
                if row[4].value: 
                    last_assignee = str(row[4].value)

            merged_title = " ".join(title_parts) # 제목 공백 머지
            merged_verification = "\n".join(verification_parts) # 검증내용 줄바꿈 머지

            # 날짜 및 완료/진행중 상태 로직
            start_date_str = ""
            end_date_str = ""
            status_period_text = raw_period

            if "~" in raw_period:
                try:
                    p_start, p_end = raw_period.split("~")
                    s_mm = int(p_start.split(".")[0])
                    e_mm = int(p_end.split(".")[0])
                    
                    start_date_str = f"{base_year}.{p_start}"
                    # 월 역전 시(예: 12.30~01.05) 종료 연도 +1
                    end_year = base_year + 1 if s_mm > e_mm else base_year
                    end_date_str = f"{end_year}.{p_end}"
                    
                    # 상태 판별: 종료 연도/월이 이번 달(2026.01)보다 작으면 완료
                    end_ym = end_year * 100 + e_mm
                    if end_ym < today_ym:
                        status_period_text = f"완료 ({raw_period})"
                    else:
                        status_period_text = f"진행중 ({raw_period})"
                except:
                    pass

            new_ws.append([merged_title, merged_verification, tc_val, status_period_text, start_date_str, end_date_str, last_assignee])

        # 3. 서식 및 테두리 적용
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in new_ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrapText=True, vertical='center')

        # 저장
        output_path = os.path.splitext(file_path)[0] + "_QA결과수정.xlsx"
        wb.save(output_path)
        messagebox.showinfo("성공", f"35~38행 포함 모든 병합이 완료되었습니다.\n경로: {output_path}")

    except Exception as e:
        messagebox.showerror("오류", f"처리 중 에러 발생: {e}")

if __name__ == "__main__":
    process_excel_v4()